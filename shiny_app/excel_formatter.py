"""
Professional Excel formatter for cross-tabulation tables
Generates market research-style tables with TOC and formatted data sheets
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_professional_excel(crosstab_report, banner_plan, output_path, study_name="Market Research Study"):
    """
    Create professional Excel output with TOC and table sheets

    Args:
        crosstab_report: Report from crosstab_engine
        banner_plan: Banner plan structure
        output_path: Path to save Excel file
        study_name: Name of the study for headers
    """

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create TOC sheet
    toc_sheet = wb.create_sheet("TOC", 0)
    _create_toc_sheet(toc_sheet, crosstab_report, study_name)

    # Create individual table sheets
    for idx, table in enumerate(crosstab_report['tables'], start=1):
        sheet_name = f"Table {idx}"
        sheet = wb.create_sheet(sheet_name)
        _create_table_sheet(sheet, table, banner_plan, idx, study_name)

    # Save workbook
    wb.save(output_path)
    return output_path


def _create_toc_sheet(sheet, report, study_name):
    """Create Table of Contents sheet"""

    # Header
    sheet['A1'] = 'Table Number'
    sheet['B1'] = 'Table Title'
    sheet['C1'] = 'Sub Title'
    sheet['D1'] = 'Base Title'

    # Style header
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for col in ['A', 'B', 'C', 'D']:
        cell = sheet[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Add table entries
    for idx, table in enumerate(report['tables'], start=1):
        row = idx + 1

        # Table number with hyperlink
        table_num_cell = sheet[f'A{row}']
        table_num_cell.value = f"Table {idx}"
        table_num_cell.hyperlink = f"#'Table {idx}'!A1"
        table_num_cell.font = Font(color="0000FF", underline="single")

        # Table title (question text)
        sheet[f'B{row}'] = table['question_text']

        # Sub title (if any special instructions)
        sub_title = ""
        if 'sub_title' in table:
            sub_title = table['sub_title']
        sheet[f'C{row}'] = sub_title

        # Base title
        sheet[f'D{row}'] = "Base: Total Respondents"

    # Adjust column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 80
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 25


def _create_table_sheet(sheet, table_data, banner_plan, table_num, study_name):
    """Create individual table sheet with professional formatting"""

    current_row = 1

    # Header: Study name
    sheet[f'A{current_row}'] = study_name
    sheet[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    # Date
    sheet[f'A{current_row}'] = datetime.now().strftime("%B, %Y")
    current_row += 1

    # Table number
    current_row += 1
    sheet[f'A{current_row}'] = f"Table {table_num}"
    sheet[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    # Question text
    sheet[f'A{current_row}'] = table_data['question_text']
    sheet[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    # Blank row
    current_row += 1

    # Base line
    sheet[f'A{current_row}'] = "Base: Total Respondents"
    current_row += 1

    # Blank row before table
    current_row += 1

    # Build banner header structure
    banner_header_row = current_row
    _write_banner_headers(sheet, banner_plan, banner_header_row)
    current_row += 3  # Banner headers take 3 rows

    # Write data rows
    stats = table_data['statistics']

    # Total row
    sheet[f'A{current_row}'] = "Total"
    col_idx = 2
    for col in stats['columns']:
        if col == 'TOTAL':
            sheet[f'{get_column_letter(col_idx)}{current_row}'] = stats['base_counts'][col]
        else:
            sheet[f'{get_column_letter(col_idx)}{current_row}'] = stats['base_counts'].get(col, 0)
        col_idx += 1
    current_row += 1

    # Response rows
    if table_data['question_type'] == 'categorical':
        _write_categorical_rows(sheet, stats, current_row)
    elif table_data['question_type'] == 'numeric':
        _write_numeric_rows(sheet, stats, current_row)
    elif table_data['question_type'] == 'likert':
        _write_likert_rows(sheet, stats, current_row)


def _write_banner_headers(sheet, banner_plan, start_row):
    """Write multi-level banner headers"""

    col_idx = 2  # Start after row labels column

    # Get all columns from all groups
    all_columns = []
    for group in banner_plan['groups']:
        for col in group['columns']:
            all_columns.append({
                'group': group['name'],
                'name': col['name'],
                'id': col.get('id', col['name'])
            })

    # Row 1: "Total" spanning first column, then H1 group names
    sheet[f'A{start_row}'] = "Total"

    current_group = None
    group_start_col = col_idx

    for col in all_columns:
        if col['group'] != current_group:
            # Write previous group header if exists
            if current_group and col_idx > group_start_col:
                sheet.merge_cells(
                    start_row=start_row,
                    start_column=group_start_col,
                    end_row=start_row,
                    end_column=col_idx - 1
                )
                sheet[f'{get_column_letter(group_start_col)}{start_row}'] = current_group
                sheet[f'{get_column_letter(group_start_col)}{start_row}'].alignment = Alignment(horizontal='center')
                sheet[f'{get_column_letter(group_start_col)}{start_row}'].font = Font(bold=True)

            current_group = col['group']
            group_start_col = col_idx

        col_idx += 1

    # Write last group
    if current_group and col_idx > group_start_col:
        sheet.merge_cells(
            start_row=start_row,
            start_column=group_start_col,
            end_row=start_row,
            end_column=col_idx - 1
        )
        sheet[f'{get_column_letter(group_start_col)}{start_row}'] = current_group
        sheet[f'{get_column_letter(group_start_col)}{start_row}'].alignment = Alignment(horizontal='center')
        sheet[f'{get_column_letter(group_start_col)}{start_row}'].font = Font(bold=True)

    # Row 2: H2 column names
    col_idx = 2
    for col in all_columns:
        cell = sheet[f'{get_column_letter(col_idx)}{start_row + 1}']
        cell.value = col['name']
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.font = Font(size=9)
        col_idx += 1

    # Row 3: "Total" labels for each column
    col_idx = 2
    for col in all_columns:
        sheet[f'{get_column_letter(col_idx)}{start_row + 2}'] = "Total"
        sheet[f'{get_column_letter(col_idx)}{start_row + 2}'].alignment = Alignment(horizontal='center')
        col_idx += 1


def _write_categorical_rows(sheet, stats, start_row):
    """Write categorical data rows with percentages"""

    current_row = start_row

    for response_value, response_data in stats['responses'].items():
        # Response label
        sheet[f'A{current_row}'] = response_value

        # Percentages for each column
        col_idx = 2
        for col_id in stats['columns']:
            pct = response_data.get(col_id, 0)
            cell = sheet[f'{get_column_letter(col_idx)}{current_row}']
            cell.value = pct / 100  # Store as decimal for percentage formatting
            cell.number_format = '0%'
            col_idx += 1

        current_row += 1

    # TOTAL MENTIONS row
    sheet[f'A{current_row}'] = "TOTAL MENTIONS"
    sheet[f'A{current_row}'].font = Font(bold=True)
    col_idx = 2
    for col_id in stats['columns']:
        cell = sheet[f'{get_column_letter(col_idx)}{current_row}']
        cell.value = stats['base_counts'].get(col_id, 0)
        col_idx += 1


def _write_numeric_rows(sheet, stats, start_row):
    """Write numeric statistics rows"""

    metrics = ['mean', 'median', 'std']
    labels = {'mean': 'Mean', 'median': 'Median', 'std': 'Std Dev'}

    current_row = start_row
    for metric in metrics:
        sheet[f'A{current_row}'] = labels[metric]

        col_idx = 2
        for col_id in stats['columns']:
            if metric in stats.get('numeric_stats', {}):
                value = stats['numeric_stats'][metric].get(col_id, 0)
                sheet[f'{get_column_letter(col_idx)}{current_row}'] = round(value, 2)
            col_idx += 1

        current_row += 1


def _write_likert_rows(sheet, stats, start_row):
    """Write likert data rows with T2B/B2B"""

    current_row = start_row

    # Regular response rows
    for response_value, response_data in stats['responses'].items():
        sheet[f'A{current_row}'] = response_value

        col_idx = 2
        for col_id in stats['columns']:
            pct = response_data.get(col_id, 0)
            cell = sheet[f'{get_column_letter(col_idx)}{current_row}']
            cell.value = pct / 100
            cell.number_format = '0%'
            col_idx += 1

        current_row += 1

    # T2B row
    if 'top_2_box' in stats:
        current_row += 1
        sheet[f'A{current_row}'] = "Top 2 Box"
        sheet[f'A{current_row}'].font = Font(bold=True)

        col_idx = 2
        for col_id in stats['columns']:
            pct = stats['top_2_box'].get(col_id, 0)
            cell = sheet[f'{get_column_letter(col_idx)}{current_row}']
            cell.value = pct / 100
            cell.number_format = '0%'
            col_idx += 1

        current_row += 1

    # B2B row
    if 'bottom_2_box' in stats:
        sheet[f'A{current_row}'] = "Bottom 2 Box"
        sheet[f'A{current_row}'].font = Font(bold=True)

        col_idx = 2
        for col_id in stats['columns']:
            pct = stats['bottom_2_box'].get(col_id, 0)
            cell = sheet[f'{get_column_letter(col_idx)}{current_row}']
            cell.value = pct / 100
            cell.number_format = '0%'
            col_idx += 1