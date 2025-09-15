#!/usr/bin/env python3
# tabplan_writer.py — polished Excel writer for Tab/Banner plans (Cue visuals)

import re
import math
from typing import Any, Dict, List, Optional
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D

# ==============================
# Brand THEMES and style factory
# (kept from your original with small tweaks)
# ==============================
def _hx(s: str) -> str:
    """Normalize hex strings for openpyxl (no leading #, uppercase)."""
    return (s or "").replace("#", "").upper()

THEMES: Dict[str, Dict[str, str]] = {
    "cue": {
        "font_name":   "Aptos",
        "font_color":  "#000000",
        "header_fill": "#212161",  # Row 9: navy
        "section_fill":"#A7B5DB",  # Screener/Main Survey: periwinkle
        "row_alt_fill":"#EEF2FB",  # slightly stronger zebra
        "border_color":"#8197D0",
        "primary":     "#FFE47A",  # Yellow top band
        "accent":      "#F2B800",  # Gold text
        "navy":        "#212161",  # Navy text
    },
    "neutral": {
        "font_name":   "Aptos",
        "font_color":  "#000000",
        "header_fill": "#E7E7E7",
        "section_fill":"#F5F5F5",
        "row_alt_fill":"#FAFAFA",
        "border_color":"#999999",
        "primary":     "#F5F5F5",
        "accent":      "#666666",
        "navy":        "#222222",
    },
}

def make_styles(theme_name: str) -> Dict[str, Any]:
    t = THEMES.get((theme_name or "").lower(), THEMES["cue"])

    thin  = Side(style="thin",  color=_hx(t["border_color"]))
    med   = Side(style="medium", color=_hx(t["border_color"]))
    thick = Side(style="thick", color=_hx(t["border_color"]))

    border_thin   = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=med, right=med, top=med, bottom=med)
    border_topbot = Border(left=thin, right=thin, top=med, bottom=med)

    fnt_normal   = Font(name=t["font_name"], size=11, color=_hx(t["font_color"]))
    fnt_bold     = Font(name=t["font_name"], size=11, bold=True, color=_hx(t["font_color"]))
    fnt_bold14   = Font(name=t["font_name"], size=14, bold=True, color=_hx(t["navy"]))
    fnt_italic   = Font(name=t["font_name"], size=11, italic=True, color=_hx(t["navy"]))
    fnt_gold     = Font(name=t["font_name"], size=11, bold=True, color=_hx(t["accent"])) # gold
    fnt_navy     = Font(name=t["font_name"], size=11, color=_hx(t["navy"]))              # navy text
    fnt_white    = Font(name=t["font_name"], size=11, bold=True, color="FFFFFF")         # white text
    fnt_mono     = Font(name="Consolas", size=10, color=_hx(t["font_color"]))

    fill_topband = PatternFill("solid", fgColor=_hx(t["primary"]))     # yellow
    fill_head    = PatternFill("solid", fgColor=_hx(t["header_fill"])) # navy
    fill_section = PatternFill("solid", fgColor=_hx(t["section_fill"]))# periwinkle
    fill_altrow  = PatternFill("solid", fgColor=_hx(t["row_alt_fill"]))
    fill_note    = PatternFill("solid", fgColor="FFFDF3")              # pale cream for notes/logic

    return {
        "BORDER_THIN": border_thin,
        "BORDER_MED":  border_medium,
        "BORDER_TOPBOT": border_topbot,
        "FNT_NORMAL": fnt_normal,
        "FNT_BOLD":   fnt_bold,
        "FNT_BOLD14": fnt_bold14,
        "FNT_ITALIC": fnt_italic,
        "FNT_GOLD":   fnt_gold,
        "FNT_NAVY":   fnt_navy,
        "FNT_WHITE":  fnt_white,
        "FNT_MONO":   fnt_mono,
        "FILL_TOPBAND": fill_topband,
        "FILL_HEAD":    fill_head,
        "FILL_SECTION": fill_section,
        "FILL_ALTROW":  fill_altrow,
        "FILL_NOTE":    fill_note,
    }

# ===================
# Column configuration
# ===================
COLS = [
    "Q#",
    "Special Question Verbiage",
    "Base Verbiage",
    "Base Definition",
    "Nets (English & code #s)",
    "Additional Table Instructions",
]
COL_WIDTHS = [12, 48, 26, 30, 44, 56]

DEFAULT_BASE_VERB = "Total (qualified respondents)"
ZEBRA = True  # set False to disable alternate-row fill

# =========
# Helpers
# =========
def is_screener(qid: str) -> bool:
    return str(qid or "").strip().upper().startswith("S")

def q_sort_key(qid: str):
    qid = str(qid or "")
    group = 0 if is_screener(qid) else 1
    m = re.match(r"^[SQ](\d+)", qid.upper())
    num = int(m.group(1)) if m else 10_000
    return (group, num, qid)

def _ws_set_header(ws, styles):
    # column widths + header row visuals
    for col_idx, (title, width) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=9, column=col_idx, value=title)
        cell.font = styles["FNT_GOLD"]      # gold text
        cell.fill = styles["FILL_HEAD"]     # navy background
        cell.border = styles["BORDER_THIN"]
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A10"
    ws.row_dimensions[9].height = 24  # taller header

def _ws_add_section(ws, row: int, title: str, styles: Dict[str, Any]) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = styles["FNT_WHITE"]       # white text
    cell.fill = styles["FILL_SECTION"]    # periwinkle
    cell.border = styles["BORDER_TOPBOT"] # thicker band
    cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[row].height = 22
    return row + 1

def _write_row(ws, row: int, values: List[Any], styles: Dict[str, Any]) -> int:
    # Only A..F have borders/fills; outside appears gridless
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=j, value=v if v is not None else "")
        c.font = styles["FNT_NORMAL"]
        c.border = styles["BORDER_THIN"]
        if ZEBRA and row > 10 and (row % 2 == 0):
            c.fill = styles["FILL_ALTROW"]
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 18
    return row + 1

def _likert_summary_instructions() -> Dict[str, str]:
    return {
        "TB":   "Show table for each statement with TB ratings data shown.",
        "T2B":  "Show table for each statement with T2B ratings data shown.",
        "B2B":  "Show table for each statement with B2B ratings data shown.",
        "BB":   "Show table for each statement with BB ratings data shown.",
        "Mean": "Show table for each statement with mean data shown.",
    }

# ---- pixel helpers for precise logo placement (no extra columns)
def _col_width_to_pixels(width: float | None) -> int:
    if width is None:
        return 64
    return int(math.floor(width * 7 + 5))

def _row_height_to_pixels(height_pts: float | None) -> int:
    if height_pts is None:
        height_pts = 15
    return int(round(height_pts * 96.0 / 72.0))

def _band_box_pixels(ws, col_start_letter: str, col_end_letter: str, row_start: int, row_end: int) -> tuple[int,int]:
    letters = [chr(c) for c in range(ord(col_start_letter), ord(col_end_letter) + 1)]
    w = sum(_col_width_to_pixels(ws.column_dimensions[L].width) for L in letters)
    h = sum(_row_height_to_pixels(ws.row_dimensions[r].height) for r in range(row_start, row_end + 1))
    return w, h

def _find_logo_path() -> Path | None:
    """
    Resolve /tab-banner-plan/ui/icons/logo.png relative to this file.
    """
    try:
        here = Path(__file__).resolve()
    except Exception:
        return None
    project_root = here.parents[1]  # .../tab-banner-plan
    candidate = project_root / "ui" / "icons" / "logo.png"
    return candidate if candidate.exists() else None

def _place_logo(ws, styles, desired_scale: float = 0.55):
    """
    Place the Cue logo in the top band (rows 1–8), right-aligned inside E..F,
    vertically centered, and explicitly scaled smaller via `desired_scale`.
    """
    logo_path = _find_logo_path()
    if not logo_path:
        return

    # Ensure the band is tall enough to center within
    for r in range(1, 9):
        ws.row_dimensions[r].height = 22  # a bit tighter than before

    # Box for the logo: columns E..F, rows 1..8 (must run AFTER header/col widths are set)
    box_w_px, box_h_px = _band_box_pixels(ws, "E", "F", 1, 8)

    # Read natural image size
    with PILImage.open(logo_path) as im:
        w0, h0 = im.size

    # Base "fit" to the band height (minus light margins)
    margin_h, margin_v = 10, 4
    fit_h = max(24, box_h_px - 2 * margin_v)
    scale_fit = fit_h / h0 if h0 else 1.0

    # Now apply explicit downscale
    scale = scale_fit * float(desired_scale)
    target_w = int(w0 * scale)
    target_h = int(h0 * scale)

    # Safety: if still wider than the available width, shrink to width
    max_w = max(40, box_w_px - 2 * margin_h)
    if target_w > max_w:
        scale = max_w / w0
        target_w = int(w0 * scale)
        target_h = int(h0 * scale)

    # Compute offsets for right-align + vertical center in the E..F box
    xoff_px = max(0, box_w_px - target_w - margin_h)
    yoff_px = max(0, (box_h_px - target_h) // 2)

    emu = 9525  # px->EMU
    start = AnchorMarker(col=4, row=0, colOff=int(xoff_px * emu), rowOff=int(yoff_px * emu))
    ext = XDRPositiveSize2D(cx=int(target_w * emu), cy=int(target_h * emu))
    anchor = OneCellAnchor(_from=start, ext=ext)

    img = XLImage(str(logo_path))
    img.width = target_w
    img.height = target_h
    img.anchor = anchor
    ws.add_image(img)

# ===============
# Main entry point
# ===============
def build_workbook(project_data: Dict[str, Any]) -> Workbook:
    proj = project_data.get("project", {}) or {}
    theme_name = (proj.get("theme") or "cue").lower()
    styles = make_styles(theme_name)

    questions = project_data.get("questions", []) or []
    questions.sort(key=lambda q: q_sort_key(q.get("id", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Tab Plan"

    # Gridless outside A:F; we only draw borders in A..F
    ws.sheet_view.showGridLines = False

    # -----------------------------
    # Top band A1:F8 — polished
    # -----------------------------
    for r in range(1, 9):
        for c in range(1, 7):  # A..F only
            cc = ws.cell(row=r, column=c)
            cc.fill = styles["FILL_TOPBAND"]
            cc.font = styles["FNT_NAVY"]

    # Title row centered
    title = (proj.get("name") or "").strip() or "Untitled Project"
    ws.merge_cells("A1:F1")
    tcell = ws["A1"]
    tcell.value = title
    tcell.font = styles["FNT_BOLD14"]
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    # Notes (rows 2–7)
    notes = [
        "",  # spacer under title
        "• Provide a banner by banner tables",
        "• Provide means and medians and stats for all numeric questions",
        "• Excel – 2 files: No freqs, just percentages. Zero decimals with a % sign.",
        "  – One file includes stat testing; one file does NOT include stat testing.",
        "• Need SPSS File",
    ]
    for idx, line in enumerate(notes, start=2):
        cell = ws.cell(row=idx, column=1, value=line)
        cell.font = styles["FNT_NAVY"]
        ws.row_dimensions[idx].height = 20

    # Column headers (row 9)
    _ws_set_header(ws, styles)

    # Logo placement (E1:F8)
    _place_logo(ws, styles)

    # Footer: left/right
    # Older openpyxl footer API fallback
    try:
        ws.oddFooter.left.text  = "Cue Insights | Confidential"
        ws.oddFooter.right.text = datetime.now().strftime("%b %d, %Y %I:%M %p")
    except Exception:
        pass  # ignore if the runtime doesn't support footers

    row = 10  # start immediately after header
    current_section = None
    instr_map = _likert_summary_instructions()

    for q in questions:
        qid = (q.get("id") or "").strip()
        qtype = (q.get("type") or "").strip().lower()
        special = q.get("special_verbiage") or ""
        base = q.get("base") or {}
        base_verb = base.get("verbiage") or DEFAULT_BASE_VERB
        base_def  = base.get("definition") or ""
        scale = q.get("scale") or {}
        labels = [l for l in (scale.get("labels") or []) if str(l).strip()]
        statements = [s for s in (q.get("statements") or []) if str(s).strip()]

        looks_likert = qtype.startswith("likert") or (statements and labels)

        tp = (q.get("exports", {}) or {}).get("tab_plan", {}) or {}
        nets_text = (tp.get("nets_text") or "").strip()
        addl_instr = (tp.get("additional_instructions") or "").strip()

        # Section rows (e.g., "Screener", "Main Survey")
        sec = "Screener" if is_screener(qid) else "Main Survey"
        if current_section != sec:
            row = _ws_add_section(ws, row, sec, styles)
            current_section = sec

        # Likert handling
        if looks_likert:
            if len(statements) <= 2:
                nets_text = nets_text or "Net: T2B, B2B"
            else:
                nets_text = nets_text or "Net: T2B, B2B"
                addl_instr = addl_instr or "Provide mean, show 1 table for each statement"

        # Main row (A..F only)
        row = _write_row(ws, row, [
            qid or "",
            special or "",
            base_verb,
            base_def,
            nets_text,
            addl_instr,
        ], styles)

        # Summary rows for 3+ statement Likerts
        if looks_likert and len(statements) >= 3:
            qprefix = (qid or "").replace(".", "_")
            for key in ("TB", "T2B", "B2B", "BB", "Mean"):
                label = f"{qprefix}_{key} Summary"
                row = _write_row(ws, row, [
                    label, "", base_verb, "", "", instr_map[key]
                ], styles)

    # 2) Banner Plan
    build_banner_sheet(wb, project_data, styles)

    return wb

# ======================
# Banner sheet utilities
# ======================
def _get_q_and_opt_label(project: dict, qid: str, code: str) -> tuple[Optional[str], Optional[str]]:
    """Finds a question and a specific option label by its ID and code."""
    questions = project.get("questions", [])
    q = next((q for q in questions if q.get("id") == qid), None)
    if not q:
        return None, None

    options = q.get("options", [])
    opt = next((o for o in options if str(o.get("code")) == str(code)), None)

    q_text = q.get("text", qid)
    opt_label = opt.get("label", code) if opt else code
    return q_text, opt_label

def _format_banner_logic(project: dict, group: dict) -> str:
    """Human-readable definition string for a banner column, without the letter prefix."""
    if not group:
        return ""

    logic_parts = []

    # Base condition from the column's reference
    ref = group.get("ref", {})
    if ref.get("qid") and ref.get("opt_id"):
        logic_parts.append(f"({ref['qid']} = {ref['opt_id']})")

    # AND conditions
    cond = group.get("cond", {})
    for clause in cond.get("all", []):
        qid = clause.get("qid")
        codes = clause.get("codes", [])
        if qid and codes:
            condition_str = f"{qid} = {codes[0]}" if len(codes) == 1 else f"{qid} in [{', '.join(map(str, codes))}]"
            logic_parts.append(f"AND ({condition_str})")

    # OR conditions
    for clause in cond.get("any", []):
        qid = clause.get("qid")
        codes = clause.get("codes", [])
        if qid and codes:
            condition_str = f"{qid} = {codes[0]}" if len(codes) == 1 else f"{qid} in [{', '.join(map(str, codes))}]"
            logic_parts.append(f"OR ({condition_str})")

    return " ".join(logic_parts)

def _letters(n: int) -> str:
    """Return (A), (B), (C)... for 1-based index n."""
    alpha = []
    x = n
    while x:
        x, r = divmod(x-1, 26)
        alpha.append(chr(65 + r))
    return f"({''.join(reversed(alpha))})"

def _bnr_write_cell(ws, r, c, v, *, font=None, fill=None, align="center", wrap=True, border=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if border: cell.border = border
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return cell

def _banner_dimensions_from_project(project: dict) -> list[dict]:
    """
    Expected structure:
      project["globals"]["banners"] = [
        { id, label, mode, dimensions: [
            { dim_id, label, source:{qid}, groups:[
                { group_id, ref:{qid,opt_id}, label_alias|null, include:bool, order:int }
            ]}
        ]}
      ]
    We’ll use the FIRST banner entry for this MVP.
    """
    g = (project or {}).get("globals", {}) or {}
    banners = (g.get("banners") or [])
    if not banners: return []
    dims = banners[0].get("dimensions") or []
    # keep only included groups and stable order
    out = []
    for d in dims:
        groups = [g for g in (d.get("groups") or []) if g.get("include", True)]
        groups.sort(key=lambda x: int(x.get("order") or 0))
        if groups:
            out.append({
                "label": (d.get("label") or "").strip() or (d.get("source") or {}).get("qid") or d.get("dim_id"),
                "groups": groups
            })
    return out

def build_banner_sheet(wb: Workbook, project: dict, styles: dict) -> None:
    """
    Create a 'Banner Plan' sheet with a professional, multi-level header structure.
    """
    ws = wb.create_sheet("Banner Plan")
    ws.sheet_view.showGridLines = False

    # --- 1) Title / header area ------------------------------------------------
    # Column sizing: A (logic) wide, B (Total) narrow, data columns medium
    ws.column_dimensions["A"].width = 64  # logic definitions
    ws.column_dimensions["B"].width = 12

    # Paint a yellow title band across a generous width
    max_cols_paint = 18
    for r in range(1, 4):
        ws.row_dimensions[r].height = 20
        for c in range(1, max_cols_paint + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = styles["FILL_TOPBAND"]
            cell.font = styles["FNT_NAVY"]

    # Title + subtitle
    proj_name = (project.get("project", {}) or {}).get("name", "") or "Untitled Project"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    _bnr_write_cell(ws, 1, 1, f"BANNERS — {proj_name}", font=styles["FNT_BOLD14"], align="left")

    _bnr_write_cell(ws, 2, 1, "90% CONFIDENCE LEVEL", font=styles["FNT_ITALIC"], align="left")

    # --- 2) Data discovery -----------------------------------------------------
    dims = _banner_dimensions_from_project(project)
    all_groups = [grp for d in dims for grp in d.get("groups", [])]

    # Set width for data columns (C onward)
    data_cols = len(all_groups)
    total_cols = 2 + data_cols
    for col_idx in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    # --- 3) Multi-level headers (Rows 4–6) ------------------------------------
    H_TITLE, H1_DIM, H2_GROUP = 4, 5, 6

    # Centered mini-title above the headers
    ws.merge_cells(start_row=H_TITLE, start_column=3, end_row=H_TITLE, end_column=total_cols)
    _bnr_write_cell(ws, H_TITLE, 3, f"{proj_name} — Banner Plan", font=styles["FNT_BOLD"], border=styles["BORDER_THIN"])

    # Total header box
    ws.merge_cells(start_row=H1_DIM, start_column=2, end_row=H2_GROUP, end_column=2)
    _bnr_write_cell(ws, H1_DIM, 2, "Total", font=styles["FNT_BOLD"], border=styles["BORDER_THIN"])

    # Dimension + group headings
    col_cursor = 3
    for d in dims:
        groups_in_dim = d.get("groups", [])
        span = len(groups_in_dim)
        if span <= 0:
            continue

        # Dimension label row (periwinkle)
        ws.merge_cells(start_row=H1_DIM, start_column=col_cursor, end_row=H1_DIM, end_column=col_cursor + span - 1)
        _bnr_write_cell(
            ws, H1_DIM, col_cursor, d["label"],
            font=styles["FNT_BOLD"], fill=styles["FILL_SECTION"], border=styles["BORDER_THIN"]
        )

        # Individual group labels with letter codes underneath
        for idx, grp in enumerate(groups_in_dim):
            _, opt_label = _get_q_and_opt_label(project, grp.get("ref", {}).get("qid"), grp.get("ref", {}).get("opt_id"))
            label = grp.get("label_alias") or opt_label or grp.get("group_id")
            letter = _letters(col_cursor - 2)  # (A), (B), ...
            cell_val = f"{label}\n{letter}"
            cell = _bnr_write_cell(ws, H2_GROUP, col_cursor, cell_val, border=styles["BORDER_THIN"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[H2_GROUP].height = 36
            col_cursor += 1

    # Bottom separator under header rows
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=H2_GROUP, column=c)
        cell.border = Border(
            left=styles["BORDER_THIN"].left,
            right=styles["BORDER_THIN"].right,
            top=styles["BORDER_THIN"].top,
            bottom=Side(style="medium", color=_hx(THEMES["cue"]["border_color"]))
        )

    # --- 4) Summary line + logic definitions ----------------------------------
    # Summary (A / B / C / …) in A5
    summary_letters = [ _letters(i+1).strip("()") for i in range(len(all_groups)) ]
    _bnr_write_cell(ws, 5, 1, " / ".join(summary_letters), font=styles["FNT_NORMAL"], align="left")

    # Logic rows start at row 8
    row = 8
    outer_top = row
    note_rows = []

    for i, grp in enumerate(all_groups):
        letter = _letters(i + 1)
        logic_str = _format_banner_logic(project, grp)
        note_rows.append(f"{letter} {logic_str}")

    for txt in note_rows:
        cell = _bnr_write_cell(ws, row, 1, txt, font=styles["FNT_MONO"], fill=styles["FILL_NOTE"],
                               align="left", wrap=True, border=styles["BORDER_THIN"])
        ws.row_dimensions[row].height = 18
        # draw empty bordered cells across the table grid so it looks like a proper sheet
        for c in range(2, total_cols + 1):
            _bnr_write_cell(ws, row, c, "", border=styles["BORDER_THIN"])
        row += 1

    outer_bottom = row - 1

    # Outer medium box around the main table (A..lastcol, header through last row)
    for r in range(H1_DIM, outer_bottom + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            left  = "medium" if c == 1 else "thin"
            right = "medium" if c == total_cols else "thin"
            top   = "medium" if r == H1_DIM else "thin"
            bot   = "medium" if r == outer_bottom else "thin"
            cell.border = Border(
                left=Side(style=left,  color=_hx(THEMES["cue"]["border_color"])),
                right=Side(style=right, color=_hx(THEMES["cue"]["border_color"])),
                top=Side(style=top,    color=_hx(THEMES["cue"]["border_color"])),
                bottom=Side(style=bot, color=_hx(THEMES["cue"]["border_color"])),
            )

    # Freeze panes: keep headers and letters visible
    ws.freeze_panes = "C8"

    # Older openpyxl footer API fallback
    try:
        ws.oddFooter.left.text  = "Cue Insights | Confidential"
        ws.oddFooter.right.text = datetime.now().strftime("%b %d, %Y %I:%M %p")
    except Exception:
        pass


